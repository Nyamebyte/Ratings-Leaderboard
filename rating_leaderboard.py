"""
This script updates an Excel file with ratings of usernames from lichess to
create a leaderboard. The Excel file will contain the names of the players.
The ratings and activity will be updated in the Excel file.
"""


import lichess
import openpyxl
from datetime import datetime


def get_rating(username, control):
    # This function returns the rating of the username
    client = lichess.Client()
    user = client.get_data(f'{username}')
    rating = user['perfs'][f'{control}']['rating']
    return rating


def get_player_list():
    # This function returns a list of usernames from the Excel file
    wb = openpyxl.load_workbook("players.xlsx")
    sheet = wb.active
    player_list = [cell.value for cell in sheet['A']]
    return player_list


def write_rating(rating_list):
    # This function prints the ratings to the Excel file
    wb = openpyxl.load_workbook("players.xlsx")
    sheet = wb.active
    for index, value in enumerate(rating_list, start=1):
        sheet.cell(row=index, column=2, value=value)
    wb.save("players.xlsx")


def check_activity(username, control):
    current_time = datetime.now().timestamp() * 1000  # Convert to milliseconds
    seven_days_ago = current_time - 7 * 24 * 60 * 60 * 1000  # Seven days in milliseconds

    myClient = lichess.Client()
    user = myClient.get_activity(f'{username}')

    for entry in user:
        interval_start = entry['interval']['start']
        if interval_start >= seven_days_ago:
            if 'games' in entry and f'{control}' in entry['games']:
                return True

    return False


def write_activity(status_list):
    wb = openpyxl.load_workbook("players.xlsx")
    sheet = wb.active
    for index, value in enumerate(status_list, start=1):
        sheet.cell(row=index, column=3, value=value)
    wb.save("players.xlsx")


def main():
    print("Specify the time control")
    print("Blitz, Bullet, Classical, Rapid, etc")
    control = input(">> ").lower()

    # Fetch the list of players
    player_list = get_player_list()

    # Update the ratings of players
    rating_list = []
    for username in player_list:
        rating = get_rating(username, control)
        rating_list.append(rating)
    write_rating(rating_list)
    print("Ratings updated!")

    # Update status activity of players
    status_list = []
    for username in player_list:
        status = check_activity(username, control)
        activity = 'Inactive'
        if status:
            activity = 'Active'
        status_list.append(activity)
    write_activity(status_list)
    print("Status updated!")


if __name__ == '__main__':
    main()

